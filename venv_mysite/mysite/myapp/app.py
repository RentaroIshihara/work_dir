import datetime
import tempfile
from flask import Flask, request, render_template
import os
import pandas as pd
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
#from datetime import datetime
#import tempfile
#import shutil
#from openpyxl.drawing.image import Image
#from werkzeug.utils import secure_filename
#from PIL import Image as PilImage

app = Flask(__name__)

# アップロードフォルダの設定
UPLOAD_FOLDER = 'uploads/'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# フォームデータを受け取るルート
@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    email = request.form['email']
    image_file = request.files['image']
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    print(f"DEBUG: Received data - Name: {name}, Email: {email}, Date: {current_date}")

    # 画像ファイルを保存
    filename = secure_filename(image_file.filename)
    image_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    image_file.save(image_path)

    # テキストファイルにデータを保存
    save_path = r'C:\Users\mi191302\work_dir\venv_mysite\mysite\form_data.txt'
    try:
        with open(save_path, 'a') as file:
            file.write(f"Name: {name}, Email: {email}, Date: {current_date}, Image: {filename}\n")
    except PermissionError as e:
        return f"PermissionError: {e}"
    except Exception as e:
        return f"Error writing to text file: {e}"

    # Excelファイルにデータを保存
    excel_path = r'C:\Users\mi191302\work_dir\venv_mysite\mysite\form_data.xlsx'
    new_data = pd.DataFrame([[name, email, current_date, filename]], columns=['Name', 'Email', 'Date', 'Image'])

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            if os.path.exists(excel_path):
                existing_data = pd.read_excel(excel_path)
                updated_data = pd.concat([existing_data, new_data], ignore_index=True)
            else:
                updated_data = new_data
            updated_data.to_excel(tmp.name, index=False)

            # 画像をExcelファイルに埋め込む
            wb = load_workbook(tmp.name)
            ws = wb.active
            img = PilImage.open(image_path)
            img.thumbnail((100, 100))
            img = Image(img)
            img.anchor = f'D{ws.max_row + 1}'
            ws.add_image(img)

            wb.save(tmp.name)
            shutil.copy(tmp.name, excel_path)
    except PermissionError as e:
        return f"PermissionError: {e}"
    except Exception as e:
        return f"Error writing to Excel file: {e}"

    return "Form submitted successfully!"

# フォームを表示するルート
@app.route('/')
def index():
    return render_template('index.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)