from flask import Flask, request, render_template, redirect, url_for
import pandas as pd
from openpyxl import load_workbook
import os

app = Flask(__name__)

# 画像の保存先ディレクトリ
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    email = request.form['email']
    image = request.files['image']

    if image:
        # 画像ファイルの保存
        image_filename = os.path.join(app.config['UPLOAD_FOLDER'], image.filename)
        image.save(image_filename)

        # 画像のURLを生成
        image_url = url_for('uploaded_file', filename=image.filename, _external=True)

        # Excelファイルのパスを設定
        excel_path = os.path.join('C://Users//renta//work_dir_4//work_dir//venv_mysite//mysite',
        '飼育表_ひな形.xlsx')
        
        # データを書き込む範囲を指定
        start_row = 2  # 書き込む開始行（例：2行目）
        name_col = 'A'  # 名前を書き込む列（例：A列）
        email_col = 'B'  # メールアドレスを書き込む列（例：B列）
        image_col = 'C'  # 画像URLを書き込む列（例：C列）

        # Excelファイルが存在しない場合は新規作成
        if not os.path.exists(excel_path):
            df = pd.DataFrame(columns=['Name', 'Email', 'Image URL'])
            df.to_excel(excel_path, index=False)

        # Excelファイルを読み込み
        wb = load_workbook(excel_path)
        ws = wb.active

        # 書き込み可能な最初の空行を探す
        row = start_row
        while ws[f'{name_col}{row}'].value is not None:
            row += 1

        # フォームデータを書き込む
        ws[f'{name_col}{row}'] = name
        ws[f'{email_col}{row}'] = email
        ws[f'{image_col}{row}'] = image_url

        # Excelファイルを保存
        wb.save(excel_path)

    return "送信できました。"

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return redirect(url_for('static', filename='uploads/' + filename), code=301)

# どこからでもip5000で接続できる。
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)